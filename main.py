import json
import openpyxl

def get_check_items():
    with open(r'extract.json', encoding='utf-8') as check:
        js_check=json.load(check)
    return js_check

def get_date(date):
    dt=date[0:10].split('-')
    return '.'.join(dt[::-1])

def int_to_float(value):
    return round((float(value)/100),2)

js_check=get_check_items()

wb=openpyxl.Workbook()
sheet=wb.active
sheet.title='Чеки'
titles=('Тип документа','Номер', 'Дата', '','','Наименование','Описание','Количество', 'Цена','Сумма')

for ind, item in enumerate(titles):
    sheet.cell(row=1,column=(ind+1)).value=item
wb.save('output.xls')
table_row=2
for item in js_check:
    check_type='Чек ККМ'
    check_no=item['ticket']['document']['receipt']['requestNumber']
    check_date=get_date(item['ticket']['document']['receipt']['dateTime'])
    for itemname in item['ticket']['document']['receipt']['items']:
        sheet.cell(row=table_row, column=1).value=check_type
        sheet.cell(row=table_row, column=2).value=check_no
        sheet.cell(row=table_row, column=3).value=check_date
        sheet.cell(row=table_row, column=4).value=''
        sheet.cell(row=table_row, column=5).value=''
        sheet.cell(row=table_row, column=6).value=itemname['name']
        sheet.cell(row=table_row, column=7).value=''
        sheet.cell(row=table_row, column=8).value=itemname['quantity']
        sheet.cell(row=table_row, column=9).value=int_to_float(itemname['price'])
        sheet.cell(row=table_row, column=10).value=int_to_float(itemname['sum'])
        table_row+=1
wb.save('output.xls')
wb.close()